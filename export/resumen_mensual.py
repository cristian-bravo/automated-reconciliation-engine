# export/resumen_mensual.py
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


HEADERS = [
    "Fecha", "Sucursal", "Caja",
    "Error Monto", "Error Doc", "Error Ambos", "Error Total",
    "Trs Realizadas", "Total Trs", "Trs Totales Día",
]

_DATE_FORMATS = ("%d_%m_%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y")


def _parse_fecha(fecha_str: str) -> datetime:
    """Convierte múltiples formatos de fecha a datetime. Retorna datetime.now() si falla."""
    if not fecha_str or str(fecha_str).upper() == "RANGO":
        return datetime.now()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(str(fecha_str).strip(), fmt)
        except ValueError:
            continue
    return datetime.now()


def _ensure_headers(ws) -> None:
    """Garantiza que la fila 1 tenga los headers con estilo bold+centrado."""
    if ws.cell(1, 1).value != "Fecha":
        ws.insert_rows(1)
        for i, h in enumerate(HEADERS, start=1):
            cell = ws.cell(1, i, h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _leer_existentes(ws) -> set:
    """
    Lee combinaciones (fecha_str, sucursal, caja) ya escritas.
    Permite que agregar_resumen_mensual sea idempotente.
    """
    existentes = set()
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 1).value
        s = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if f and s and c:
            if isinstance(f, datetime):
                f = f.strftime("%d/%m/%Y")
            existentes.add((str(f).strip(), str(s).strip(), str(c).strip()))
    return existentes


def agregar_resumen_mensual(path: str, fecha: str, local: str, resumen: dict) -> None:
    """
    Agrega filas al resumen mensual garantizando unicidad por (Fecha, Sucursal, Caja).
    Si la combinación ya existe, la omite → sin duplicados incluso con re-ejecuciones.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)

    if not os.path.exists(path):
        wb = Workbook()
        wb.save(path)

    wb = load_workbook(path)
    ws = wb.active
    _ensure_headers(ws)

    fecha_excel = _parse_fecha(fecha)
    fecha_key = fecha_excel.strftime("%d/%m/%Y")
    existentes = _leer_existentes(ws)
    row = ws.max_row + 1

    for caja, d in resumen["por_caja"].items():
        clave = (fecha_key, str(local).strip(), str(caja).strip())
        if clave in existentes:
            continue  # Deduplicación: no insertar si ya existe

        ws.cell(row, 1, fecha_excel).number_format = "DD/MM/YYYY"
        ws.cell(row, 2, local)
        ws.cell(row, 3, caja)
        ws.cell(row, 4, int(d["Error Monto"]))
        ws.cell(row, 5, int(d["Error Doc"]))
        ws.cell(row, 6, int(d["Error Ambos"]))
        ws.cell(row, 7, int(d["Error Total"]))
        ws.cell(row, 8, int(d["Trs Realizadas"]))
        ws.cell(row, 9, round(float(d["Total Trs"]), 2))
        ws.cell(row, 10, round(float(resumen["total_trs"]), 2))
        row += 1

    wb.save(path)
