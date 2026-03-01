# export/resumen_diario.py
import re
from openpyxl import load_workbook


def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def _col(ws, name):
    headers = [_norm(c.value) for c in ws[1]]
    try:
        return headers.index(_norm(name)) + 1
    except ValueError:
        return None


def _caja(v):
    m = re.search(r"caja\s*[-_ ]*(\d+)", _norm(v))
    return f"CAJA {m.group(1)}" if m else None


def generar_resumen_diario(path):
    wb = load_workbook(path)
    ws = wb.active

    c_valor = _col(ws, "valor")
    c_estado = _col(ws, "estado")
    c_rec = _col(ws, "recaudador") or _col(ws, "caja")
    c_det = _col(ws, "detalle")
    c_fecha = _col(ws, "fecha") # o FECHA

    if None in (c_valor, c_estado, c_rec, c_det):
        print("⚠️ Advertencia: No se generó el resumen, faltan columnas clave en el Excel.")
        return {}

    # Estructura: { fecha: { "por_caja": {...}, "total_trs": 0.0, "total_trs_count": 0 } }
    resumen_por_fecha = {}

    for r in range(2, ws.max_row + 1):
        if all(ws.cell(r, c).value in (None, "") for c in range(1, ws.max_column + 1)):
            continue

        valor = float(ws.cell(r, c_valor).value or 0)
        estado = _norm(ws.cell(r, c_estado).value).upper()
        caja = _caja(ws.cell(r, c_rec).value)
        
        fecha_eval = _norm(ws.cell(r, c_fecha).value).upper() if c_fecha else "RANGO"
        if not fecha_eval:
            fecha_eval = "RANGO"

        if not caja:
            continue
            
        if fecha_eval not in resumen_por_fecha:
            resumen_por_fecha[fecha_eval] = {
                "por_caja": {},
                "total_trs": 0.0,
                "total_trs_count": 0
            }

        padre_fecha = resumen_por_fecha[fecha_eval]
        
        padre_fecha["total_trs"] += valor
        padre_fecha["total_trs_count"] += 1

        padre_fecha["por_caja"].setdefault(caja, {
            "Error Monto": 0,
            "Error Doc": 0,
            "Error Ambos": 0,
            "Error Total": 0,
            "Trs Realizadas": 0,
            "Total Trs": 0.0,
        })

        # ✅ SIEMPRE contar la transferencia
        padre_fecha["por_caja"][caja]["Trs Realizadas"] += 1
        padre_fecha["por_caja"][caja]["Total Trs"] += valor

        # ❌ Solo si NO es verde, es error
        if estado == "COINCIDE":
            continue

        detalle = _norm(ws.cell(r, c_det).value)

        if estado not in ("COINCIDE", "REVISAR"):
            padre_fecha["por_caja"][caja]["Error Ambos"] += 1
        elif "monto coincide" in detalle:
            padre_fecha["por_caja"][caja]["Error Doc"] += 1
        elif "documento coincide" in detalle:
            padre_fecha["por_caja"][caja]["Error Monto"] += 1
        else:
            padre_fecha["por_caja"][caja]["Error Ambos"] += 1

    # Computando error total por cada fecha
    for fecha, datos in resumen_por_fecha.items():
        for caja in datos["por_caja"]:
            datos["por_caja"][caja]["Error Total"] = (
                datos["por_caja"][caja]["Error Monto"]
                + datos["por_caja"][caja]["Error Doc"]
                + datos["por_caja"][caja]["Error Ambos"]
            )
        # Ordenando caja
        datos["por_caja"] = dict(sorted(datos["por_caja"].items(), key=lambda x: int(x[0].split()[-1])))
        datos["total_trs"] = round(datos["total_trs"], 2)

    # Retorna un diccionario indexado por fecha
    return resumen_por_fecha
