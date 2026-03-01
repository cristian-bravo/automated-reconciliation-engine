# export/excel.py
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

from config import CONCIL_DIR, LOCALES, LOCALES_DISPLAY, OUTPUT_DIR
from export.resumen_diario import generar_resumen_diario
from export.resumen_mensual import agregar_resumen_mensual
from core.utils import clean_illegal_chars
from core.logger import log_info


# ---------------------------------------------------------------------
# Utils
# ---------------------------------------------------------------------
def _drop_last_4_rows_df(df):
    if df is None or len(df.index) <= 4:
        return df.iloc[0:0].copy()
    return df.iloc[:-4].copy()


def _insert_header_rows(path, row1_text, row2_text):
    wb = load_workbook(path)
    ws = wb.active
    ws.insert_rows(1, 2)
    ws["A1"] = row1_text
    ws["A2"] = row2_text
    
    max_col = ws.max_column
    if max_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        
    for r in [1, 2]:
        cell = ws.cell(r, 1)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    wb.save(path)

def _renombrar_columnas(df):
    renames = {
        "Factura": "Fact.",
        "Recaudador": "Caja",
        "Nro. Documento": "N° Doc."
    }
    return df.rename(columns=renames)

def _mapear_nombres_locales(df):
    """Reemplaza los nombres completos de locales por sus alias cortos de reporte."""
    if "local_nombre" not in df.columns:
        return df
    df["local_nombre"] = df["local_nombre"].apply(
        lambda x: LOCALES_DISPLAY.get(str(x).strip().lower(), x)
    )
    return df


def _delete_column_by_header(ws, header_row_idx, header_name):
    headers = [c.value for c in ws[header_row_idx]]
    if header_name in headers:
        ws.delete_cols(headers.index(header_name) + 1)


def _fecha_token_dd_mm_yyyy(value):
    """
    Convierte una fecha (str/Timestamp) al token dd_mm_yyyy.
    Retorna None si no puede parsearse.
    """
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.strftime("%d_%m_%Y")

    raw = str(value).strip()
    if not raw:
        return None

    raw_norm = raw.replace("_", "/").replace(".", "/").replace("-", "/")
    dt = pd.to_datetime(raw_norm, dayfirst=True, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.strftime("%d_%m_%Y")


def _nombre_consolidado_rango_desde_trs(trs_all, fallback_fecha="rango"):
    """
    Genera conciliacion_consolidado_<min>__<max>.xlsx usando las fechas reales
    presentes en el consolidado ya calculado (sin recalcular lógica).
    """
    if trs_all is None or trs_all.empty or "Fecha" not in trs_all.columns:
        return f"conciliacion_consolidado_{fallback_fecha}.xlsx"

    fechas = []
    for value in trs_all["Fecha"]:
        token = _fecha_token_dd_mm_yyyy(value)
        if token:
            fechas.append(token)

    if not fechas:
        return f"conciliacion_consolidado_{fallback_fecha}.xlsx"

    dts = pd.Series(pd.to_datetime([f.replace("_", "/") for f in fechas], dayfirst=True, errors="coerce")).dropna()
    if dts.empty:
        return f"conciliacion_consolidado_{fallback_fecha}.xlsx"

    fecha_min = dts.min().strftime("%d_%m_%Y")
    fecha_max = dts.max().strftime("%d_%m_%Y")
    return f"conciliacion_consolidado_{fecha_min}__{fecha_max}.xlsx"


def _token_rango_desde_trs(trs_all, fallback_fecha="rango"):
    nombre = _nombre_consolidado_rango_desde_trs(trs_all, fallback_fecha=fallback_fecha)
    prefix = "conciliacion_consolidado_"
    suffix = ".xlsx"
    if nombre.startswith(prefix) and nombre.endswith(suffix):
        return nombre[len(prefix):-len(suffix)]
    return fallback_fecha


def _resolver_directorio_exportacion(fecha, trs_all, es_rango=False):
    if es_rango:
        rango_token = _token_rango_desde_trs(trs_all, fallback_fecha=fecha)
        out_dir = os.path.join(OUTPUT_DIR, "por_rango", rango_token)
    else:
        out_dir = os.path.join(OUTPUT_DIR, "por_dia", fecha)

    os.makedirs(out_dir, exist_ok=True)
    return out_dir


# ---------------------------------------------------------------------
# Pintado por estado
# ---------------------------------------------------------------------
def pintar_excel(path, header_row=1, data_start_row=2):
    wb = load_workbook(path)
    ws = wb.active

    fills = {
        "COINCIDE": PatternFill("solid", fgColor="C6EFCE"),
        "REVISAR": PatternFill("solid", fgColor="FFEB9C"),
        "DEFAULT": PatternFill("solid", fgColor="FFC7CE"),
    }

    headers = [c.value for c in ws[header_row]]
    if "estado" not in headers:
        wb.save(path)
        return

    estado_col = headers.index("estado") + 1

    for r in range(data_start_row, ws.max_row + 1):
        estado = ws.cell(r, estado_col).value
        fill = fills.get(estado, fills["DEFAULT"])
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = fill

    wb.save(path)


# ---------------------------------------------------------------------
# Presentación final del Excel diario
# ---------------------------------------------------------------------
def _finalize_excel_presentation(path, local_name):
    # 1) Pintar
    pintar_excel(path)

    # 2) Borrar columnas técnicas
    wb = load_workbook(path)
    ws = wb.active
    for col in ["local_code", "local_nombre", "Beneficiario", "estado"]:
        _delete_column_by_header(ws, 1, col)
    wb.save(path)

    # 3) Borrar últimas 4 filas
    wb = load_workbook(path)
    ws = wb.active
    for _ in range(4):
        if ws.max_row > 1:
            ws.delete_rows(ws.max_row)
    wb.save(path)

    # 4) Insertar títulos
    _insert_header_rows(path, "DETALLES DE TRANSFERENCIAS", local_name)


# ---------------------------------------------------------------------
# Bloque RESUMEN DIARIO (tabla nueva)
# ---------------------------------------------------------------------
def escribir_resumen_diario(path, resumen, local_name, fecha):
    wb = load_workbook(path)
    ws = wb.active

    headers = [
        "Recaudador",
        "Error Monto",
        "Error Doc",
        "Error Ambos",
        "Error Total",
        "Trs Realizadas",
        "Total Trs",
    ]

    start_row = ws.max_row + 2

    ws.cell(start_row, 1, f"Resumen General {local_name} {fecha}")
    max_c = len(headers)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_c)
    title_cell = ws.cell(start_row, 1)
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(headers, start=1):
        ws.cell(start_row + 1, i, h)

    row = start_row + 2

    totales = {
        "Error Monto": 0,
        "Error Doc": 0,
        "Error Ambos": 0,
        "Error Total": 0,
        "Trs Realizadas": 0,
        "Total Trs": 0.0,
    }

    for caja, d in resumen["por_caja"].items():
        ws.cell(row, 1, caja)
        ws.cell(row, 2, d["Error Monto"])
        ws.cell(row, 3, d["Error Doc"])
        ws.cell(row, 4, d["Error Ambos"])
        ws.cell(row, 5, d["Error Total"])
        ws.cell(row, 6, d["Trs Realizadas"])
        ws.cell(row, 7, round(d["Total Trs"], 2))

        for k in totales:
            totales[k] += d[k]

        row += 1

    # Fila TOTAL
    ws.cell(row, 1, "Total")
    ws.cell(row, 2, totales["Error Monto"])
    ws.cell(row, 3, totales["Error Doc"])
    ws.cell(row, 4, totales["Error Ambos"])
    ws.cell(row, 5, totales["Error Total"])
    ws.cell(row, 6, totales["Trs Realizadas"])
    ws.cell(row, 7, round(totales["Total Trs"], 2))

    wb.save(path)


# ---------------------------------------------------------------------
# Export principal
# ---------------------------------------------------------------------
def exportar_excel(fecha, mov, trs_por_local, trs_all, es_rango=False):
    os.makedirs(CONCIL_DIR, exist_ok=True)
    out_dir = _resolver_directorio_exportacion(fecha, trs_all, es_rango=es_rango)

    # ================= MOVIMIENTOS =================
    mov_out = os.path.join(out_dir, f"conciliacion_movimientos_{fecha}.xlsx")
    mov_df = _drop_last_4_rows_df(mov.copy())
    mov_df = mov_df.map(clean_illegal_chars)
    mov_df = _renombrar_columnas(mov_df)
    mov_df.to_excel(mov_out, index=False)
    _finalize_excel_presentation(mov_out, "MOVIMIENTOS BANCARIOS")

    # ================= TRS POR LOCAL =================
    for local_code in trs_por_local:
        df_out = trs_all[trs_all["local_code"] == local_code].copy()
        df_out = _drop_last_4_rows_df(df_out)
        df_out = df_out.map(clean_illegal_chars)
        df_out = _renombrar_columnas(df_out)

        out = os.path.join(out_dir, f"conciliacion_trs_{local_code}_{fecha}.xlsx")
        df_out.to_excel(out, index=False)


        # 1) Resumen con archivo CRUDO
        resumen_por_fecha = generar_resumen_diario(out)

        # 2) Presentación
        local_name = LOCALES.get(local_code, local_code)
        _finalize_excel_presentation(out, local_name)

        # 3 y 4) Insertar bloque de resumen al final del Excel local y en el Mensual para CADA fecha hallada
        if isinstance(resumen_por_fecha, dict):
            for fecha_k, resumen_k in resumen_por_fecha.items():
                
                # Agregarlo como una subtabla al final del documento local
                escribir_resumen_diario(out, resumen_k, local_name, fecha_k)

                # Exportarlo a la base de datos histórica correspondiente
                output_path = os.path.join(OUTPUT_DIR, "resumen_mensual_rango.xlsx") if es_rango else os.path.join(OUTPUT_DIR, "resumen_mensual.xlsx")
                
                agregar_resumen_mensual(
                    path=output_path,
                    fecha=fecha_k,
                    local=local_name,
                    resumen=resumen_k,
                )

    # ================= CONSOLIDADO (diario / rango) =================
    if es_rango:
        nombre_consolidado = _nombre_consolidado_rango_desde_trs(trs_all, fallback_fecha=fecha)
    else:
        nombre_consolidado = f"conciliacion_consolidado_{fecha}.xlsx"

    exportar_consolidado_rango(trs_all, nombre_archivo=nombre_consolidado, output_dir=out_dir)

    return mov_out


def exportar_consolidado_rango(trs_all, nombre_archivo=None, output_dir=None):
    """
    Exporta un sexto archivo que consolida todos los movimientos locales
    junto con su resultado de conciliación.
    """
    if not nombre_archivo:
        nombre_archivo = _nombre_consolidado_rango_desde_trs(trs_all, fallback_fecha="sin_fechas")

    if output_dir is None:
        output_dir = CONCIL_DIR

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, nombre_archivo)
    
    # Preparamos el DF para exportar (borrar colas y limpiar chars)
    df_out = _drop_last_4_rows_df(trs_all.copy())
    df_out = df_out.map(clean_illegal_chars)
    
    # Orden o campos base (Fecha, Local, Info del movimiento, Resultado)
    # Ordenar por Fecha cronológicamente si es posible
    if "Fecha" in df_out.columns:
        # Respaldar la columna original
        fecha_orig = df_out["Fecha"]
        # Intentar convertir (coercionando errores a NaT para evitar AssertionError en formatos mezclados)
        fecha_dt = pd.to_datetime(df_out["Fecha"], format="%d/%m/%Y", errors="coerce")
        # Si falló completamente (todo NaT) o no, rellenamos con una base
        df_out["_Fecha_sort"] = fecha_dt
        df_out = df_out.sort_values(by=["_Fecha_sort", "local_nombre"])
        df_out = df_out.drop(columns=["_Fecha_sort"], errors="ignore")
        # Restaurar la presentación string para que no salga 'NaT' si algo falló
        df_out["Fecha"] = fecha_dt.dt.strftime("%d/%m/%Y").fillna(fecha_orig)
        
    df_out = _mapear_nombres_locales(df_out)
    # Renombrar local_nombre -> Local para presentación
    df_out = df_out.rename(columns={"local_nombre": "Local"})
    df_out = _renombrar_columnas(df_out)
    
    # Seleccionaremos las columnas importantes usando las variables consolidadas (renombradas)
    cols_requeridas = ["Fecha", "Fact.", "N\u00b0 Doc.", "Valor", "Local", "Caja", "Detalle", "estado"]
    
    cols_existentes = [c for c in cols_requeridas if c in df_out.columns]
    
    # Exportar Data
    df_out[cols_existentes].to_excel(out_path, index=False)
    
    # Dar presentación (Pintar estado)
    _finalize_excel_presentation(out_path, "RESUMEN CONSOLIDADO DE LOCALES")
    
    log_info(f"\U0001F9FE Generado: {os.path.basename(out_path)}")
