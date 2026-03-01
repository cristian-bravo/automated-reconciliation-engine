# export/pdf.py
import os
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors

def excel_a_pdf(xlsx_path, pdf_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    data = []
    fills = []

    for row in ws.iter_rows():
        row_data, row_fill = [], []
        for cell in row:
            row_data.append("" if cell.value is None else str(cell.value))
            row_fill.append(cell.fill.fgColor.rgb if cell.fill else None)
        data.append(row_data)
        fills.append(row_fill)

    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))
    table = Table(data, repeatRows=1)

    style = TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.black),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
    ])

    for r in range(1, len(data)):
        color = fills[r][0]
        if color == "FFC6EFCE":
            bg = colors.lightgreen
        elif color == "FFFFEB9C":
            bg = colors.khaki
        elif color == "FFFFC7CE":
            bg = colors.salmon
        else:
            continue
        style.add("BACKGROUND", (0, r), (-1, r), bg)

    table.setStyle(style)
    doc.build([table])

def exportar_pdf(mov_excel_path):
    if not os.path.exists(mov_excel_path):
        return None
    pdf_path = mov_excel_path.replace(".xlsx", ".pdf")
    excel_a_pdf(mov_excel_path, pdf_path)
    return pdf_path
